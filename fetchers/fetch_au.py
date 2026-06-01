"""Fetch Australian name data from NSW BDM (1952–2025) + VIC BDM (2008–2025).

Also opportunistically reads manually-downloaded QLD + SA CSV files if present
under data/raw/au/qld/*.csv and data/raw/au/sa/*.csv — both states publish to
CKAN portals behind CloudFront WAFs that block scripted downloads, so the user
must drop the files in via a browser. See README in those folders for URLs.
"""
import csv
import io
import json
import time
import urllib.request
from pathlib import Path
import openpyxl
from _common import RAW_DIR, write_normalized

AU_RAW = RAW_DIR / 'au'
UA = {'User-Agent': 'Mozilla/5.0'}

NSW_URL = "https://data.nsw.gov.au/data/dataset/a677cbe2-91e1-4e45-b771-08830d3d9e41/resource/2adcb228-9101-4c95-a786-b3216539b4a2/download/popular_baby_names_1952_to_2025.csv"
NSW_FILE = AU_RAW / 'nsw_1952_2025.csv'

VIC_CKAN = "https://discover.data.vic.gov.au/api/3/action/package_show?id=popular-baby-names"


def download(url: str, dest: Path):
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 5_000:
        return
    print(f"  [AU] downloading {dest.name}")
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=60) as r, dest.open('wb') as f:
        f.write(r.read())
    time.sleep(1)


def nsw_rows():
    download(NSW_URL, NSW_FILE)
    with NSW_FILE.open(encoding='utf-8-sig') as f:
        r = csv.DictReader(f)
        for row in r:
            try:
                year = int(row['Year'])
                count = int(row['Number'])
                sex = row['Gender']
                name = row['Name']
            except (KeyError, ValueError, TypeError):
                continue
            if count > 0 and name:
                yield (year, sex, name, count)


def vic_resources() -> list[tuple[int, str]]:
    req = urllib.request.Request(VIC_CKAN, headers=UA)
    with urllib.request.urlopen(req, timeout=30) as r:
        d = json.loads(r.read().decode('utf-8'))
    out = []
    for res in d['result']['resources']:
        if res.get('format', '').upper() != 'XLSX':
            continue
        name = res.get('name', '')
        # Name looks like 'Popular Baby Names 2024'
        for tok in name.split():
            if tok.isdigit() and 2000 <= int(tok) <= 2030:
                out.append((int(tok), res['url']))
                break
    return sorted(set(out))


def vic_rows():
    for year, url in vic_resources():
        dest = AU_RAW / f"vic{year}.xlsx"
        try:
            download(url, dest)
        except Exception as e:
            print(f"  [AU] skip vic{year}: {e}")
            continue
        wb = openpyxl.load_workbook(str(dest), data_only=True)
        ws = wb[wb.sheetnames[0]]
        # Layout: Rank | Name-Male | Count | Rank | Name-Female | Count
        for r in range(2, ws.max_row + 1):
            row = [c.value for c in ws[r]]
            if len(row) < 6:
                continue
            # Males
            if isinstance(row[1], str) and isinstance(row[2], (int, float)) and row[2] > 0:
                yield (year, 'M', row[1].strip(), int(row[2]))
            # Females
            if isinstance(row[4], str) and isinstance(row[5], (int, float)) and row[5] > 0:
                yield (year, 'F', row[4].strip(), int(row[5]))


def manual_state_rows(state: str):
    """Read any CSVs the user dropped under data/raw/au/<state>/.

    Tolerant of column-name variation across QLD, SA, and other state portals.
    Each CSV must contain identifiable Name / Count / Sex / Year columns.
    """
    folder = AU_RAW / state
    if not folder.exists():
        return
    found = 0
    for path in sorted(folder.glob('*.csv')):
        try:
            with path.open(encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                cols = {c.lower(): c for c in (reader.fieldnames or [])}
                col = lambda *names: next((cols[n] for n in names if n in cols), None)
                c_year = col('year', 'birth year')
                c_name = col('name', 'first name', 'firstname', 'given name')
                c_count = col('number', 'count', 'frequency', 'total')
                c_sex = col('gender', 'sex')
                if not (c_year and c_name and c_count and c_sex):
                    print(f"  [AU/{state}] skip {path.name}: missing columns")
                    continue
                n = 0
                for row in reader:
                    try:
                        year = int(row[c_year])
                        count = int(row[c_count])
                        name = row[c_name]
                        sex = row[c_sex]
                    except (KeyError, ValueError, TypeError):
                        continue
                    if count > 0 and name:
                        yield (year, sex, name, count)
                        n += 1
                found += n
                print(f"  [AU/{state}] {path.name}: {n:,} rows")
        except Exception as e:
            print(f"  [AU/{state}] error {path.name}: {e}")
    if not found:
        print(f"  [AU/{state}] no data found (drop CSVs into data/raw/au/{state}/)")


def main():
    print("[AU] NSW BDM 1952–2025 + VIC BDM 2008–2025 (+ QLD/SA if present)")
    AU_RAW.mkdir(parents=True, exist_ok=True)
    rows = list(nsw_rows())
    print(f"  [AU] NSW: {len(rows):,} rows")
    vrows = list(vic_rows())
    print(f"  [AU] VIC: {len(vrows):,} rows")
    rows.extend(vrows)
    rows.extend(manual_state_rows('qld'))
    rows.extend(manual_state_rows('sa'))
    write_normalized('AU', rows)


if __name__ == '__main__':
    main()
