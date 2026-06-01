"""Fetch UK (England & Wales) name data from ONS.

Two sources:
  1. Adhoc consolidated 1996-2016 .xls (all names, not just top 100)
  2. Annual top-100 .xls / .xlsx files for 2017-2024
"""
import time
import urllib.request
from pathlib import Path
import xlrd
import openpyxl
from _common import RAW_DIR, write_normalized

UK_RAW = RAW_DIR / 'uk'

UA = {'User-Agent': 'Mozilla/5.0'}

ADHOC = {
    'F': "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2016/adhocallbabynames1996to2016.xls",
    'M': "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2016/adhocallbabynames1996to2016.xls",
}

# Annual top-100 files we need to extend coverage past 2016.
# (year, sex, url, format) — format is 'xls' or 'xlsx'.
ANNUAL = [
    (2017, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2017/2017girlsnames.xls", 'xls'),
    (2017, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2017/2017boysnames.xls", 'xls'),
    (2018, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2018/2018girlsnames.xls", 'xls'),
    (2018, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2018/2018boysnames.xls", 'xls'),
    (2019, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2019/2019girlsnames.xlsx", 'xlsx'),
    (2019, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2019/2019boysnames.xlsx", 'xlsx'),
    (2020, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2020/2020girlsnames.xlsx", 'xlsx'),
    (2020, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2020/2020boysnames.xlsx", 'xlsx'),
    (2021, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2021/2021girlsnames.xlsx", 'xlsx'),
    (2021, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2021/2021boysnamesupdated1.xlsx", 'xlsx'),
    (2022, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2022/girlsnames2022.xlsx", 'xlsx'),
    (2022, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2022/boysnames2022.xlsx", 'xlsx'),
    (2023, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2023/girlsnames2023.xlsx", 'xlsx'),
    (2023, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2023/boysnames2023.xlsx", 'xlsx'),
    (2024, 'F', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsgirls/2024/girlsnames2024.xlsx", 'xlsx'),
    (2024, 'M', "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/livebirths/datasets/babynamesenglandandwalesbabynamesstatisticsboys/2024/boysnames2024.xlsx", 'xlsx'),
]


def download(url: str, dest: Path):
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 10_000:
        return
    # ONS rate-limits aggressive downloads with 429s. Polite delay between fetches.
    for attempt in range(4):
        try:
            print(f"  [UK] downloading {dest.name}")
            req = urllib.request.Request(url, headers=UA)
            with urllib.request.urlopen(req, timeout=60) as r, dest.open('wb') as f:
                f.write(r.read())
            time.sleep(3)
            return
        except urllib.error.HTTPError as e:
            if e.code == 429 and attempt < 3:
                wait = 10 * (attempt + 1)
                print(f"  [UK] 429 rate-limited, waiting {wait}s")
                time.sleep(wait)
                continue
            raise


def adhoc_rows():
    """Yield rows from the 1996-2016 consolidated files (all names, not just top 100)."""
    for sex, url in ADHOC.items():
        sheet_name = 'Girls' if sex == 'F' else 'Boys'
        dest = UK_RAW / f"adhoc-{sheet_name.lower()}-1996-2016.xls"
        download(url, dest)
        wb = xlrd.open_workbook(str(dest))
        ws = wb.sheet_by_name(sheet_name)
        # Row 4 (0-indexed): year headers in columns 1,3,5,...
        # Row 5: Name | Rank | Count | Rank | Count | ...
        year_row = ws.row_values(4)
        years_at_col: dict[int, int] = {}
        for c, v in enumerate(year_row):
            if isinstance(v, (int, float)) and 1990 <= v <= 2020:
                years_at_col[c] = int(v)
        for r in range(6, ws.nrows):
            name = ws.cell_value(r, 0)
            if not isinstance(name, str) or not name.strip():
                continue
            for c, year in years_at_col.items():
                # year col c has Rank; count is at c+1
                count = ws.cell_value(r, c + 1)
                if isinstance(count, (int, float)) and count > 0:
                    yield (year, sex, name, int(count))


def _scan_triples(rows_2d):
    """Heuristic: yield (name, count) for every (int_rank, str_name, int_count) triple
    found anywhere in a 2D sheet. Robust to header variations and wide 1-50 / 51-100 layouts."""
    for row in rows_2d:
        n = len(row)
        for c in range(n - 2):
            a, b, d = row[c], row[c + 1], row[c + 2]
            if (isinstance(a, (int, float)) and 1 <= a <= 200
                    and isinstance(b, str) and b.strip()
                    and isinstance(d, (int, float)) and d > 0):
                yield b.strip(), int(d)


def annual_rows():
    """Yield rows from individual annual top-100 files (2017-2024)."""
    for year, sex, url, fmt in ANNUAL:
        dest = UK_RAW / f"{year}{'girls' if sex == 'F' else 'boys'}.{fmt}"
        try:
            download(url, dest)
        except Exception as e:
            print(f"  [UK] skip {dest.name}: {e}")
            continue
        if fmt == 'xlsx':
            wb = openpyxl.load_workbook(str(dest), data_only=True)
            tname = next((s for s in wb.sheetnames
                          if s.lower().replace(' ', '_').startswith('table_1')
                          or s.strip() == '1'), None)
            if not tname:
                continue
            ws = wb[tname]
            rows_2d = [[c.value for c in row] for row in ws.iter_rows()]
        else:  # .xls
            wb = xlrd.open_workbook(str(dest))
            sn = next((s for s in wb.sheet_names()
                       if s.lower().replace(' ', '_').startswith('table_1')
                       or s.strip() == '1'), None)
            if not sn:
                continue
            ws = wb.sheet_by_name(sn)
            rows_2d = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        seen = set()
        for name, count in _scan_triples(rows_2d):
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            yield (year, sex, name, count)


def main():
    print("[UK] ONS England & Wales 1996–2024")
    rows = list(adhoc_rows())
    rows.extend(annual_rows())
    write_normalized('GB', rows)


if __name__ == '__main__':
    main()
